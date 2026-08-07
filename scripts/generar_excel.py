#!/usr/bin/env python3
"""Genera hojas de Excel con formato de despacho a partir de datos tabulares.

Para informes al cliente, cuadres, colas de trabajo, listados de facturas o
cualquier tabla que haya que entregar en .xlsx en vez de en texto.

Aplica solo el formato que un despacho necesita: cabecera legible, anchos
ajustados, importes en formato español con separador de miles, fechas en
DD/MM/AAAA, fila de totales y paneles inmovilizados.

Entrada
-------
CSV (con `;` o `,`) o JSON. El JSON admite dos formas:

    [ {"cliente": "...", "importe": 1234.56}, ... ]          una hoja

    {"hojas": {                                              varias hojas
        "Resumen":  [ {...}, {...} ],
        "Detalle":  [ {...} ]
    }}

Uso
---
    python3 scripts/generar_excel.py --datos informe.json --salida informe.xlsx \\
        --titulo "Cierre 2T/2026" --totales importe,cuota

    python3 scripts/generar_excel.py --datos facturas.csv --salida facturas.xlsx \\
        --importes base,cuota,total --fechas fecha

Las columnas de importe se detectan solas si se llaman base, cuota, importe,
total, retencion, percepcion o similar; con --importes se fuerza la lista.

Requiere openpyxl:  pip install openpyxl
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from lib.registro import a_decimal  # noqa: E402

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    print("Falta openpyxl. Instalalo con: pip install openpyxl", file=sys.stderr)
    raise SystemExit(2)

# Formato de numero español: miles con punto, decimales con coma.
FORMATO_IMPORTE = '#,##0.00\\ "€"'
FORMATO_FECHA = "DD/MM/YYYY"

PISTAS_IMPORTE = re.compile(
    r"(importe|base|cuota|total|retenc|percep|ingreso|gasto|saldo|euros?|deducib|"
    r"devengad|facturad|estadistic)", re.I)
PISTAS_FECHA = re.compile(r"(fecha|vencimiento|devengo|alta|baja|limite)", re.I)

AZUL = "1F4E79"
GRIS = "F2F2F2"


def leer(ruta: Path) -> dict[str, list[dict]]:
    if ruta.suffix.lower() == ".json":
        datos = json.loads(ruta.read_text(encoding="utf-8"))
        if isinstance(datos, dict) and "hojas" in datos:
            return {str(k): list(v) for k, v in datos["hojas"].items()}
        if isinstance(datos, list):
            return {"Datos": datos}
        raise SystemExit("El JSON debe ser una lista de objetos o {'hojas': {...}}")
    with ruta.open(encoding="utf-8-sig", newline="") as f:
        muestra = f.read(4096)
        f.seek(0)
        try:
            dialecto = csv.Sniffer().sniff(muestra, delimiters=";,\t")
        except csv.Error:
            dialecto = csv.excel
            dialecto.delimiter = ";"
        return {"Datos": [dict(fila) for fila in csv.DictReader(f, dialect=dialecto)]}


def convertir(valor, es_importe: bool, es_fecha: bool):
    """Deja el valor en el tipo que Excel entiende, para que sume y ordene bien."""
    if valor is None or valor == "":
        return None
    if es_importe:
        try:
            return float(a_decimal(valor))
        except Exception:
            return valor
    if es_fecha:
        if isinstance(valor, (date, datetime)):
            return valor
        for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(str(valor).strip(), formato).date()
            except ValueError:
                continue
        return valor
    if isinstance(valor, (int, float, Decimal, date, datetime)):
        return valor
    return str(valor)


def escribir_hoja(ws, filas: list[dict], titulo: str | None,
                  importes: set[str], fechas: set[str], totales: set[str]) -> None:
    if not filas:
        ws["A1"] = "Sin datos"
        return

    columnas = list(dict.fromkeys(k for fila in filas for k in fila))
    fila_cab = 1

    if titulo:
        ws.cell(1, 1, titulo).font = Font(bold=True, size=13, color=AZUL)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(columnas), 1))
        fila_cab = 3

    borde = Border(bottom=Side(style="thin", color="BFBFBF"))
    for col, nombre in enumerate(columnas, start=1):
        c = ws.cell(fila_cab, col, nombre.replace("_", " ").capitalize())
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, fila in enumerate(filas):
        for col, nombre in enumerate(columnas, start=1):
            es_imp, es_fec = nombre in importes, nombre in fechas
            c = ws.cell(fila_cab + 1 + i, col, convertir(fila.get(nombre), es_imp, es_fec))
            c.border = borde
            if es_imp:
                c.number_format = FORMATO_IMPORTE
                c.alignment = Alignment(horizontal="right")
            elif es_fec:
                c.number_format = FORMATO_FECHA
                c.alignment = Alignment(horizontal="center")
            if i % 2:
                c.fill = PatternFill("solid", fgColor=GRIS)

    if totales:
        f = fila_cab + 1 + len(filas)
        ws.cell(f, 1, "TOTAL").font = Font(bold=True)
        for col, nombre in enumerate(columnas, start=1):
            if nombre not in totales:
                continue
            letra = get_column_letter(col)
            c = ws.cell(f, col, f"=SUM({letra}{fila_cab + 1}:{letra}{f - 1})")
            c.font = Font(bold=True)
            c.number_format = FORMATO_IMPORTE
            c.border = Border(top=Side(style="double"))

    for col, nombre in enumerate(columnas, start=1):
        largo = max([len(str(nombre))] +
                    [len(str(fila.get(nombre, ""))) for fila in filas[:200]])
        ws.column_dimensions[get_column_letter(col)].width = min(max(largo + 3, 11), 46)

    ws.freeze_panes = ws.cell(fila_cab + 1, 1)
    ws.auto_filter.ref = (f"A{fila_cab}:"
                          f"{get_column_letter(len(columnas))}{fila_cab + len(filas)}")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--datos", required=True, type=Path, help="CSV o JSON de entrada")
    ap.add_argument("--salida", required=True, type=Path)
    ap.add_argument("--titulo", help="Título en la primera fila de cada hoja")
    ap.add_argument("--importes", default="", help="Columnas de importe, separadas por coma")
    ap.add_argument("--fechas", default="", help="Columnas de fecha, separadas por coma")
    ap.add_argument("--totales", default="", help="Columnas que llevan fila de total")
    ap.add_argument("--sin-deteccion", action="store_true",
                    help="No deducir importes ni fechas por el nombre de la columna")
    args = ap.parse_args()

    if not args.datos.exists():
        print(f"No existe {args.datos}", file=sys.stderr)
        return 2

    hojas = leer(args.datos)
    if not any(hojas.values()):
        print("El fichero de entrada no tiene filas", file=sys.stderr)
        return 2

    manual_imp = {c.strip() for c in args.importes.split(",") if c.strip()}
    manual_fec = {c.strip() for c in args.fechas.split(",") if c.strip()}
    totales = {c.strip() for c in args.totales.split(",") if c.strip()} or manual_imp

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for nombre_hoja, filas in hojas.items():
        columnas = list(dict.fromkeys(k for fila in filas for k in fila))
        importes = set(manual_imp)
        fechas = set(manual_fec)
        if not args.sin_deteccion:
            importes |= {c for c in columnas if PISTAS_IMPORTE.search(c)}
            fechas |= {c for c in columnas if PISTAS_FECHA.search(c)}
        fechas -= importes
        ws = wb.create_sheet(str(nombre_hoja)[:31])
        escribir_hoja(ws, filas, args.titulo, importes, fechas, totales & set(columnas))

    args.salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.salida)

    print(f"Excel generado: {args.salida}")
    for nombre_hoja, filas in hojas.items():
        print(f"  {nombre_hoja}: {len(filas)} filas")
    print("\nRevisa el resultado antes de enviarlo al cliente.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
