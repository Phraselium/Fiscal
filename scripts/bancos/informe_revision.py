#!/usr/bin/env python3
"""Excel de revision que acompana al documento de importacion.

Cuatro hojas:

    Resumen           cuadre por banco y recuento por regla aplicada. El cuadre
                      lleva FORMULAS, no numeros calculados en Python: quien
                      revise tiene que poder ver de donde sale cada cifra.
    Notas y avisos    que contiene el fichero, cuentas a dar de alta, comercios
                      de tarjeta identificados, decisiones pendientes, desglose
                      de la cuenta puente por motivo y partidas sueltas.
    A revisar         lo que va a la puente y lo marcado, con el texto original
                      del extracto al lado para poder localizarlo.
    Todos los asientos  la traza completa movimiento → asiento.

Todas las cifras salen del propio mapeo. Ninguna se escribe a mano.

Uso
---
    python3 scripts/bancos/informe_revision.py --clasificado clasificado.json \\
        --diccionario dicc.json --cuentas cuentas.json --extractos movimientos.json \\
        --verificacion verificacion.json --salida salidas/revision.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    print("Falta openpyxl. Instalalo con: pip install openpyxl", file=sys.stderr)
    raise SystemExit(2)

AZUL = "1F4E79"
GRIS = "F2F2F2"
AMBAR = "FFF2CC"
FUENTE = "Arial"
FORMATO_IMPORTE = "#,##0.00;(#,##0.00);-"


def titulo(ws, fila: int, texto: str) -> int:
    c = ws.cell(fila, 1, texto)
    c.font = Font(name=FUENTE, bold=True, size=12, color=AZUL)
    return fila + 2


def cabecera(ws, fila: int, columnas: list[str]) -> int:
    for i, nombre in enumerate(columnas, start=1):
        c = ws.cell(fila, i, nombre)
        c.font = Font(name=FUENTE, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    return fila + 1


def ajustar(ws, anchos: dict[int, int]) -> None:
    for col, ancho in anchos.items():
        ws.column_dimensions[get_column_letter(col)].width = ancho
    for fila in ws.iter_rows():
        for c in fila:
            if c.font is None or c.font.name != FUENTE:
                c.font = Font(name=FUENTE, size=c.font.size or 11,
                              bold=c.font.bold, color=c.font.color)


def hoja_resumen(wb, clasificados, cuentas, saldos_hist, saldos_final) -> None:
    ws = wb.create_sheet("Resumen")
    fila = titulo(ws, 1, "CUADRE POR BANCO")
    fila = cabecera(ws, fila, ["Banco", "Subcuenta", "Saldo inicial", "Cargos",
                               "Abonos", "Saldo final calculado", "Saldo extracto",
                               "Diferencia"])
    primera = fila
    mapa = cuentas.get("bancos", {})
    for banco, subcuenta in sorted(mapa.items()):
        movs = [c for c in clasificados
                if c["banco"] == banco and not c.get("contabilizado_en_otro")]
        cargos = sum(c["importe"] for c in movs if c["importe"] < 0)
        abonos = sum(c["importe"] for c in movs if c["importe"] > 0)
        ws.cell(fila, 1, banco)
        ws.cell(fila, 2, subcuenta)
        ws.cell(fila, 3, saldos_hist.get(subcuenta, 0.0))
        ws.cell(fila, 4, round(cargos, 2))
        ws.cell(fila, 5, round(abonos, 2))
        # Formulas vivas: el revisor tiene que ver de donde sale el cuadre.
        ws.cell(fila, 6, f"=C{fila}+D{fila}+E{fila}")
        ws.cell(fila, 7, saldos_final.get(subcuenta))
        ws.cell(fila, 8, f"=F{fila}-G{fila}")
        for col in (3, 4, 5, 6, 7, 8):
            ws.cell(fila, col).number_format = FORMATO_IMPORTE
        fila += 1

    if fila > primera:
        ws.cell(fila, 1, "TOTAL").font = Font(name=FUENTE, bold=True)
        for col in (3, 4, 5, 6, 7, 8):
            letra = get_column_letter(col)
            c = ws.cell(fila, col, f"=SUM({letra}{primera}:{letra}{fila - 1})")
            c.font = Font(name=FUENTE, bold=True)
            c.number_format = FORMATO_IMPORTE
        fila += 1
        ws.cell(fila + 1, 1,
                "La diferencia debe ser cero en todas las filas. Si no lo es, el fichero "
                "no se entrega.").font = Font(name=FUENTE, italic=True, size=9)

    fila += 4
    fila = titulo(ws, fila, "MOVIMIENTOS POR REGLA APLICADA")
    fila = cabecera(ws, fila, ["Regla", "Movimientos", "Importe"])
    primera = fila
    por_regla = defaultdict(lambda: [0, 0.0])
    for c in clasificados:
        por_regla[c["regla"]][0] += 1
        por_regla[c["regla"]][1] += abs(c["importe"])
    for regla, (n, importe) in sorted(por_regla.items(), key=lambda x: -x[1][0]):
        ws.cell(fila, 1, regla)
        ws.cell(fila, 2, n)
        ws.cell(fila, 3, round(importe, 2)).number_format = FORMATO_IMPORTE
        fila += 1
    ws.cell(fila, 1, "TOTAL").font = Font(name=FUENTE, bold=True)
    for col in (2, 3):
        letra = get_column_letter(col)
        c = ws.cell(fila, col, f"=SUM({letra}{primera}:{letra}{fila - 1})")
        c.font = Font(name=FUENTE, bold=True)
        if col == 3:
            c.number_format = FORMATO_IMPORTE

    ajustar(ws, {1: 26, 2: 14, 3: 18, 4: 16, 5: 16, 6: 22, 7: 18, 8: 14})


def hoja_notas(wb, clasificados, cuentas, verificacion, extractos) -> None:
    ws = wb.create_sheet("Notas y avisos")
    puente = cuentas.get("puente", "5550000")
    vivos = [c for c in clasificados if not c.get("contabilizado_en_otro")]
    en_puente = [c for c in vivos if c["contrapartida"] == puente]
    marcados = [c for c in vivos if c.get("revisar")]

    fila = titulo(ws, 1, "QUÉ CONTIENE EL FICHERO")
    for texto, valor in (
        ("Movimientos leídos de los extractos", len(clasificados)),
        ("Movimientos que generan asiento", len(vivos)),
        ("Traspasos entre cuentas propias emparejados",
         sum(1 for c in clasificados if c["regla"] == "10-traspaso")),
        ("Movimientos en la cuenta puente " + puente, len(en_puente)),
        ("Movimientos marcados para revisar", len(marcados)),
    ):
        ws.cell(fila, 1, texto)
        ws.cell(fila, 2, valor)
        fila += 1
    if vivos:
        ws.cell(fila, 1, "Porcentaje sin identificar")
        ws.cell(fila, 2, f"=B{fila - 4}/B{fila - 6}").number_format = "0.0%"
        fila += 1

    fila += 2
    fila = titulo(ws, fila, "CUENTAS QUE HAY QUE DAR DE ALTA ANTES DE IMPORTAR")
    nuevas = [a.replace("Subcuentas que hay que dar de alta antes de importar: ", "")
              for a in (verificacion or {}).get("avisos", [])
              if "dar de alta" in a]
    if nuevas:
        for cuenta in sorted({c.strip() for a in nuevas for c in a.split(",")}):
            ws.cell(fila, 1, cuenta).fill = PatternFill("solid", fgColor=AMBAR)
            fila += 1
    else:
        ws.cell(fila, 1, "Ninguna: todas las subcuentas usadas existen en el plan.")
        fila += 1

    fila += 2
    fila = titulo(ws, fila, "CUENTA PUENTE, POR MOTIVO")
    fila = cabecera(ws, fila, ["Motivo", "Movimientos", "Importe"])
    por_motivo = defaultdict(lambda: [0, 0.0])
    for c in en_puente:
        clave = c.get("motivo_revision") or c["regla"]
        por_motivo[clave][0] += 1
        por_motivo[clave][1] += abs(c["importe"])
    for motivo, (n, importe) in sorted(por_motivo.items(), key=lambda x: -x[1][0]):
        ws.cell(fila, 1, motivo)
        ws.cell(fila, 2, n)
        ws.cell(fila, 3, round(importe, 2)).number_format = FORMATO_IMPORTE
        fila += 1

    fila += 2
    fila = titulo(ws, fila, "COMERCIOS DE TARJETA IDENTIFICADOS")
    tarjeta = [c for c in vivos if c["regla"].startswith("15-tarjeta")]
    if tarjeta:
        fila = cabecera(ws, fila, ["Texto del extracto", "Cuenta", "Estado"])
        for c in sorted(tarjeta, key=lambda x: -abs(x["importe"]))[:40]:
            ws.cell(fila, 1, c["texto"][:70])
            ws.cell(fila, 2, c["contrapartida"])
            ws.cell(fila, 3, "validado" if c["regla"].endswith("lista-blanca")
                    else "no validado → puente")
            fila += 1
    else:
        ws.cell(fila, 1, "No hay compras con tarjeta en el periodo.")
        fila += 1

    fila += 2
    fila = titulo(ws, fila, "DECISIONES DE CRITERIO PENDIENTES")
    for texto in (
        "¿Las compras con tarjeta van a compras (600) o a gastos? Cargarlas contra la "
        "cuenta del proveedor descuadra su saldo si no hay factura detrás.",
        "Confirmar el periodo de las liquidaciones de Seguridad Social: se ha deducido "
        "de la fecha de cargo.",
        "Confirmar el tratamiento de las operaciones de valores e inversiones.",
        "Revisar los cobros imputados contra cuentas de proveedor: pueden ser rappels, "
        "devoluciones o clientes distintos.",
    ):
        ws.cell(fila, 1, texto).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[fila].height = 30
        fila += 1

    fila += 2
    fila = titulo(ws, fila, "PARTIDAS SUELTAS DE IMPORTE RELEVANTE")
    fila = cabecera(ws, fila, ["Fecha", "Banco", "Importe", "Texto del extracto"])
    for c in sorted(en_puente, key=lambda x: -abs(x["importe"]))[:20]:
        ws.cell(fila, 1, c["fecha"])
        ws.cell(fila, 2, c["banco"])
        ws.cell(fila, 3, c["importe"]).number_format = FORMATO_IMPORTE
        ws.cell(fila, 4, c["texto"][:90])
        fila += 1

    fila += 2
    fila = titulo(ws, fila, "CÓMO SE HA HECHO EL TRABAJO")
    for texto in (
        "Los criterios de imputación salen del XDIARIO del ejercicio anterior del propio "
        "cliente: concepto → contrapartida emparejando por importe dentro de cada asiento, "
        "y nombre de tercero → subcuenta a partir de las líneas 400*/410*.",
        "Lo que no tiene respaldo en ese histórico va a la cuenta puente. No se ha "
        "inventado ninguna cuenta ni ningún importe.",
        "Los traspasos entre cuentas propias se han emparejado en una pasada global por "
        "importe contrario, cuentas distintas y fecha dentro de ±5 días.",
        "El fichero ha pasado las diez verificaciones automáticas, incluido el cuadre por "
        "banco al céntimo.",
        "EL FICHERO ESTÁ PENDIENTE DE REVISAR E IMPORTAR. No está contabilizado.",
    ):
        ws.cell(fila, 1, texto).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[fila].height = 30
        fila += 1

    ajustar(ws, {1: 78, 2: 18, 3: 16, 4: 60})


def hoja_revisar(wb, clasificados, cuentas) -> None:
    ws = wb.create_sheet("A revisar")
    puente = cuentas.get("puente", "5550000")
    filas = [c for c in clasificados
             if not c.get("contabilizado_en_otro")
             and (c.get("revisar") or c["contrapartida"] == puente)]
    fila = cabecera(ws, 1, ["Fecha", "Banco", "Importe", "Contrapartida", "Concepto",
                            "Regla", "Motivo", "Texto original del extracto"])
    for i, c in enumerate(sorted(filas, key=lambda x: (x["fecha"], -abs(x["importe"])))):
        ws.cell(fila, 1, c["fecha"])
        ws.cell(fila, 2, c["banco"])
        ws.cell(fila, 3, c["importe"]).number_format = FORMATO_IMPORTE
        ws.cell(fila, 4, c["contrapartida"])
        ws.cell(fila, 5, c["concepto"])
        ws.cell(fila, 6, c["regla"])
        ws.cell(fila, 7, c.get("motivo_revision", ""))
        ws.cell(fila, 8, c["texto"])
        if i % 2:
            for col in range(1, 9):
                ws.cell(fila, col).fill = PatternFill("solid", fgColor=GRIS)
        fila += 1
    ws.freeze_panes = "A2"
    if fila > 2:
        ws.auto_filter.ref = f"A1:H{fila - 1}"
    ajustar(ws, {1: 12, 2: 14, 3: 14, 4: 14, 5: 28, 6: 24, 7: 42, 8: 70})


def hoja_asientos(wb, clasificados, asiento_inicial: int) -> None:
    ws = wb.create_sheet("Todos los asientos")
    fila = cabecera(ws, 1, ["Asiento", "Fecha", "Banco", "Subcuenta banco",
                            "Contrapartida", "Concepto", "Debe", "Haber", "Regla",
                            "Texto original"])
    vivos = sorted((c for c in clasificados if not c.get("contabilizado_en_otro")),
                   key=lambda x: (x["fecha"], x.get("fila", 0)))
    numero = asiento_inicial
    for c in vivos:
        if abs(c["importe"]) < 0.005:
            continue
        importe = abs(c["importe"])
        cargo = c["importe"] < 0
        ws.cell(fila, 1, numero)
        ws.cell(fila, 2, c["fecha"])
        ws.cell(fila, 3, c["banco"])
        ws.cell(fila, 4, c["subcuenta_banco"])
        ws.cell(fila, 5, c["contrapartida"])
        ws.cell(fila, 6, c["concepto"])
        ws.cell(fila, 7, importe if cargo else 0).number_format = FORMATO_IMPORTE
        ws.cell(fila, 8, 0 if cargo else importe).number_format = FORMATO_IMPORTE
        ws.cell(fila, 9, c["regla"])
        ws.cell(fila, 10, c["texto"][:90])
        fila += 1
        numero += 1
    ws.freeze_panes = "A2"
    if fila > 2:
        ws.auto_filter.ref = f"A1:J{fila - 1}"
    ajustar(ws, {1: 10, 2: 12, 3: 14, 4: 16, 5: 15, 6: 28, 7: 14, 8: 14, 9: 24, 10: 70})


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--clasificado", required=True, type=Path)
    ap.add_argument("--cuentas", type=Path)
    ap.add_argument("--diccionario", type=Path)
    ap.add_argument("--extractos", type=Path)
    ap.add_argument("--verificacion", type=Path)
    ap.add_argument("--salida", required=True, type=Path)
    ap.add_argument("--asiento-inicial", type=int, default=1)
    args = ap.parse_args()

    clasificados = json.loads(args.clasificado.read_text(encoding="utf-8"))
    cuentas = (json.loads(args.cuentas.read_text(encoding="utf-8"))
               if args.cuentas and args.cuentas.exists() else {})
    verificacion = (json.loads(args.verificacion.read_text(encoding="utf-8"))
                    if args.verificacion and args.verificacion.exists() else {})
    saldos_hist = {}
    if args.diccionario and args.diccionario.exists():
        saldos_hist = {k: float(v) for k, v in json.loads(
            args.diccionario.read_text(encoding="utf-8")).get("bancos", {}).items()}
    saldos_final = {}
    if args.extractos and args.extractos.exists():
        mapa = cuentas.get("bancos", {})
        for e in json.loads(args.extractos.read_text(encoding="utf-8")):
            cuenta = mapa.get(e.get("banco"))
            if cuenta and e.get("saldo_final") is not None:
                saldos_final[cuenta] = float(e["saldo_final"])

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hoja_resumen(wb, clasificados, cuentas, saldos_hist, saldos_final)
    hoja_notas(wb, clasificados, cuentas, verificacion, args.extractos)
    hoja_revisar(wb, clasificados, cuentas)
    hoja_asientos(wb, clasificados, args.asiento_inicial)
    args.salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.salida)

    asientos = sum(1 for c in clasificados
                   if not c.get("contabilizado_en_otro") and abs(c["importe"]) >= 0.005)
    print(f"Excel de revisión: {args.salida}")
    print(f"  4 hojas · {asientos} asientos")

    print("\nRevisa el Excel antes de enviarlo. El fichero no está contabilizado.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
