#!/usr/bin/env python3
"""Lee extractos bancarios de varios bancos y los normaliza.

Cada banco exporta distinto, y la fila de cabecera cambia entre versiones del
mismo banco. Por eso NO se asume un numero de fila fijo: se busca la primera fila
que contenga a la vez algo parecido a fecha, concepto e importe.

Formatos contrastados (la cabecera es orientativa, se localiza sola):

    Banco       Cabecera   Columnas de interes
    Ibercaja    fila 7     Fecha Oper, Concepto + Descripcion, Importe, Saldo
    BBVA        fila 16    F. CONTABLE, CONCEPTO + BENEFICIARIO + OBSERVACIONES, IMPORTE, SALDO
    Santander   fila 8     Fecha Operacion, Concepto, Importe, Saldo
    Sabadell    fila 9     F. Operativa, Concepto, Importe, Saldo   (.xls, requiere xlrd)

Para anadir un banco basta con ampliar PISTAS: no hay logica por banco.

Salida normalizada, un registro por movimiento:
    banco · fecha · texto (mayusculas sin acentos) · importe (float con signo) · saldo

Uso
---
    python3 scripts/bancos/leer_extractos.py extracto1.xlsx extracto2.xls --json salida.json
    python3 scripts/bancos/leer_extractos.py carpeta_extractos/ --resumen
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import unicodedata
from dataclasses import dataclass, asdict
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from lib.registro import a_decimal  # noqa: E402

# Sinonimos de cada columna. Se comparan normalizados (sin acentos, en mayusculas).
PISTAS = {
    "fecha": ["FECHA OPER", "F CONTABLE", "FECHA OPERACION", "F OPERATIVA", "FECHA VALOR",
              "FECHA CONTABLE", "F VALOR", "FECHA"],
    "texto": ["CONCEPTO", "DESCRIPCION", "BENEFICIARIO", "ORDENANTE", "OBSERVACIONES",
              "BENEFICIARIO/ORDENANTE", "DETALLE", "MOVIMIENTO", "REFERENCIA"],
    "importe": ["IMPORTE", "IMPORTE EUR", "CARGO/ABONO", "IMPORTE OPERACION"],
    "saldo": ["SALDO", "SALDO EUR", "SALDO POSTERIOR"],
}

# Bancos reconocibles por el texto de las primeras filas del fichero.
FIRMAS_BANCO = {
    "IBERCAJA": "Ibercaja", "BBVA": "BBVA", "BANCO BILBAO": "BBVA",
    "SANTANDER": "Santander", "SABADELL": "Sabadell", "CAIXA": "CaixaBank",
    "BANKINTER": "Bankinter", "UNICAJA": "Unicaja", "ABANCA": "Abanca",
    "KUTXA": "Kutxabank", "CAJAMAR": "Cajamar", "OPENBANK": "Openbank",
}

RE_IBAN = re.compile(r"\bES\d{2}[ -]?(?:\d{4}[ -]?){5}\b", re.I)
RE_CUENTA = re.compile(r"\b\d{4}[ -]?\d{4}[ -]?\d{2}[ -]?\d{10}\b")


def normalizar(texto) -> str:
    if texto is None:
        return ""
    t = unicodedata.normalize("NFD", str(texto).upper())
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", t).strip()


def normalizar_cabecera(texto) -> str:
    """Como normalizar, pero sin puntuacion: «F. CONTABLE» y «F CONTABLE» son lo mismo."""
    return re.sub(r"\s+", " ", re.sub(r"[.·:;,_/-]", " ", normalizar(texto))).strip()


@dataclass
class Movimiento:
    banco: str
    cuenta: str
    fecha: date
    texto: str
    importe: float
    saldo: float | None
    fichero: str
    fila: int

    def como_dict(self) -> dict:
        d = asdict(self)
        d["fecha"] = self.fecha.isoformat()
        return d


@dataclass
class Extracto:
    banco: str
    cuenta: str
    titular: str
    fichero: str
    movimientos: list[Movimiento]

    @property
    def saldo_inicial(self) -> float | None:
        """Saldo antes del primer movimiento: saldo de la primera fila − su importe."""
        if not self.movimientos:
            return None
        primero = min(self.movimientos, key=lambda m: (m.fecha, m.fila))
        if primero.saldo is None:
            return None
        return round(primero.saldo - primero.importe, 2)

    @property
    def saldo_final(self) -> float | None:
        if not self.movimientos:
            return None
        ultimo = max(self.movimientos, key=lambda m: (m.fecha, m.fila))
        return ultimo.saldo


# --------------------------------------------------------------------------
# Lectura de la rejilla, sea cual sea el formato del fichero
# --------------------------------------------------------------------------

def rejilla(ruta: Path) -> list[list]:
    sufijo = ruta.suffix.lower()
    if sufijo == ".csv":
        with ruta.open(encoding="utf-8-sig", errors="replace", newline="") as f:
            muestra = f.read(8192)
            f.seek(0)
            try:
                dialecto = csv.Sniffer().sniff(muestra, delimiters=";,\t|")
            except csv.Error:
                dialecto = csv.excel
                dialecto.delimiter = ";"
            return [list(fila) for fila in csv.reader(f, dialect=dialecto)]

    if sufijo in (".xlsx", ".xlsm"):
        try:
            import openpyxl
        except ImportError:
            raise SystemExit("Para leer .xlsx hace falta openpyxl: pip install openpyxl")
        import warnings
        warnings.filterwarnings("ignore")
        wb = openpyxl.load_workbook(ruta, data_only=True)
        return [list(f) for f in wb[wb.sheetnames[0]].iter_rows(values_only=True)]

    if sufijo == ".xls":
        try:
            import xlrd
        except ImportError:
            raise SystemExit(
                f"{ruta.name} es .xls antiguo y hace falta xlrd: pip install 'xlrd==1.2.0'.\n"
                "Alternativa: abrelo y guardalo como .xlsx.")
        libro = xlrd.open_workbook(str(ruta))
        hoja = libro.sheet_by_index(0)
        return [[hoja.cell_value(f, c) for c in range(hoja.ncols)] for f in range(hoja.nrows)]

    raise SystemExit(f"No se leer {ruta.name}: usa .xlsx, .xls o .csv")


def localizar_cabecera(filas: list[list]) -> tuple[int, dict[str, list[int]]]:
    """Primera fila que tenga a la vez fecha, concepto e importe."""
    for numero, fila in enumerate(filas[:60]):
        celdas = [normalizar_cabecera(c) for c in fila]
        encontrado: dict[str, list[int]] = {}
        for clave, pistas in PISTAS.items():
            for col, celda in enumerate(celdas):
                if not celda:
                    continue
                if any(celda == p or celda.startswith(p) for p in pistas):
                    encontrado.setdefault(clave, []).append(col)
        if {"fecha", "texto", "importe"} <= set(encontrado):
            return numero, encontrado
    raise SystemExit(
        "No encuentro la fila de cabecera: necesito columnas de fecha, concepto e importe.\n"
        "Si el banco usa otros nombres, añádelos a PISTAS en leer_extractos.py.")


def a_fecha(valor) -> date | None:
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, (int, float)):  # serie de Excel
        from datetime import timedelta
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(valor))).date()
        except (ValueError, OverflowError):
            return None
    texto = str(valor).strip()[:10]
    for formato in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d.%m.%Y", "%d-%m-%y", "%d/%m/%y"):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    return None


def a_importe(valor) -> float | None:
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    try:
        return float(a_decimal(str(valor)))
    except Exception:
        return None


def cabecera_del_fichero(filas: list[list], hasta: int) -> tuple[str, str, str]:
    """Banco, cuenta y titular de las primeras filas, antes de la tabla."""
    texto = " ".join(normalizar(c) for fila in filas[:hasta] for c in fila if c)
    banco = next((v for k, v in FIRMAS_BANCO.items() if k in texto), "")
    cuenta = ""
    if (m := RE_IBAN.search(texto)) or (m := RE_CUENTA.search(texto)):
        cuenta = re.sub(r"[ -]", "", m.group(0)).upper()
    titular = ""
    for fila in filas[:hasta]:
        for celda in fila:
            n = normalizar(celda)
            if any(p in n for p in ("TITULAR", "CLIENTE", "NOMBRE")) and len(n) > 12:
                titular = re.sub(r"^(TITULAR|CLIENTE|NOMBRE)[:\s]*", "", n).strip()
                break
        if titular:
            break
    return banco, cuenta, titular


def leer_extracto(ruta: Path, banco_forzado: str = "") -> Extracto:
    filas = rejilla(ruta)
    if not filas:
        raise SystemExit(f"{ruta.name} está vacío")
    fila_cab, columnas = localizar_cabecera(filas)
    banco, cuenta, titular = cabecera_del_fichero(filas, fila_cab)
    banco = banco_forzado or banco or ruta.stem[:20]

    movimientos: list[Movimiento] = []
    for numero, fila in enumerate(filas[fila_cab + 1:], start=fila_cab + 2):
        def celda(indice):
            return fila[indice] if indice < len(fila) else None

        fecha = next((f for f in (a_fecha(celda(c)) for c in columnas["fecha"]) if f), None)
        if fecha is None:
            continue
        importe = next((i for i in (a_importe(celda(c)) for c in columnas["importe"])
                        if i is not None), None)
        if importe is None:
            continue

        partes = [normalizar(celda(c)) for c in columnas["texto"]]
        texto = " ".join(p for p in partes if p)
        if not texto:
            continue

        saldo = next((s for s in (a_importe(celda(c)) for c in columnas.get("saldo", []))
                      if s is not None), None)
        movimientos.append(Movimiento(banco, cuenta, fecha, texto,
                                      round(importe, 2),
                                      round(saldo, 2) if saldo is not None else None,
                                      ruta.name, numero))

    movimientos.sort(key=lambda m: (m.fecha, m.fila))
    return Extracto(banco, cuenta, titular, ruta.name, movimientos)


def leer_todos(rutas: list[Path]) -> list[Extracto]:
    ficheros: list[Path] = []
    for ruta in rutas:
        if ruta.is_dir():
            ficheros.extend(sorted(p for p in ruta.iterdir()
                                   if p.suffix.lower() in (".xlsx", ".xls", ".csv")))
        else:
            ficheros.append(ruta)
    if not ficheros:
        raise SystemExit("No hay extractos que leer")
    return [leer_extracto(f) for f in ficheros]


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("extractos", nargs="+", type=Path)
    ap.add_argument("--json", type=Path, help="Vuelca los movimientos normalizados")
    ap.add_argument("--resumen", action="store_true")
    args = ap.parse_args()

    extractos = leer_todos(args.extractos)

    print(f"{len(extractos)} extractos\n")
    for e in extractos:
        cargos = sum(m.importe for m in e.movimientos if m.importe < 0)
        abonos = sum(m.importe for m in e.movimientos if m.importe > 0)
        print(f"  {e.banco:<12} {e.fichero}")
        print(f"      cuenta {e.cuenta or '(no localizada)':<26} "
              f"{len(e.movimientos)} movimientos")
        if e.movimientos:
            print(f"      del {e.movimientos[0].fecha:%d/%m/%Y} al "
                  f"{e.movimientos[-1].fecha:%d/%m/%Y}")
            print(f"      saldo inicial {e.saldo_inicial}   cargos {cargos:.2f}   "
                  f"abonos {abonos:.2f}   saldo final {e.saldo_final}")
            if e.saldo_inicial is not None and e.saldo_final is not None:
                calculado = round(e.saldo_inicial + cargos + abonos, 2)
                if abs(calculado - e.saldo_final) > 0.01:
                    print(f"      ⚠ el extracto no cuadra: calculado {calculado}, "
                          f"declarado {e.saldo_final}. Faltan movimientos.")
        if args.resumen:
            for m in e.movimientos[:5]:
                print(f"        {m.fecha:%d/%m/%y} {m.importe:>11.2f}  {m.texto[:58]}")

    if args.json:
        args.json.parent.mkdir(parents=True, exist_ok=True)
        args.json.write_text(json.dumps(
            [{"banco": e.banco, "cuenta": e.cuenta, "titular": e.titular,
              "fichero": e.fichero, "saldo_inicial": e.saldo_inicial,
              "saldo_final": e.saldo_final,
              "movimientos": [m.como_dict() for m in e.movimientos]}
             for e in extractos], ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"\nEscrito {args.json}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
