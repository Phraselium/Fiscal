#!/usr/bin/env python3
"""Lee el fichero de control del despacho y responde sin gastar contexto.

El Control.xlsx es el sistema operativo del despacho: la matriz cliente x modelo
con el estado de cada presentacion. Este script lo consulta y devuelve colas de
trabajo ya filtradas, en vez de volcar 85 filas x 25 columnas al modelo.

Regla de uso para el asistente: NUNCA leas el Excel entero para responder "que
falta". Ejecuta el subcomando que corresponda y trabaja sobre su salida.

Subcomandos
-----------
resumen     KPIs del periodo: % presentado, pendientes, vencidos, proximos.
cola        Cola de trabajo filtrada por estado, modelo, responsable o cliente.
cliente     Ficha completa de un cliente: todos sus modelos y su estado.
modelo      Estado de un modelo en toda la cartera.
vencimientos Calendario con dias restantes y semaforo.
alquileres  Meses pendientes de facturar.
huecos      Incoherencias y datos que faltan (auditoria del propio control).
marcar      Cambia el estado de una celda (cliente + modelo).
exportar    Vuelca la matriz a CSV o JSON para procesar en otra herramienta.

Ejemplos
--------
    python3 scripts/control.py resumen --fichero Control.xlsx
    python3 scripts/control.py cola --estado Revisar
    python3 scripts/control.py cola --estado Pendiente --modelo 303
    python3 scripts/control.py cliente "EJEMPLO CLIENTE SL"
    python3 scripts/control.py huecos
    python3 scripts/control.py marcar --cliente "EJEMPLO CLIENTE SL" --modelo 303 --estado Presentado --marca env
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    print("Falta openpyxl. Instalalo con: pip install openpyxl", file=sys.stderr)
    raise SystemExit(2)

# Flujo de trabajo declarado en la hoja Introduccion del control.
FLUJO = [
    "Sin dato",
    "Documentación",
    "Pendiente",
    "Revisar",
    "Presentado",
    "Liquidación pendiente",
]
FUERA_DE_FLUJO = ["No aplica", "Baja"]
ESTADOS = FLUJO + FUERA_DE_FLUJO

# Estados que representan trabajo vivo, en orden de urgencia para el despacho.
ACCIONABLES = ["Revisar", "Pendiente", "Documentación", "Sin dato"]

MODELOS_TRIMESTRALES = {"111", "115", "123", "130", "303", "349", "202"}
MODELOS_ANUALES = {"200/24", "347", "180", "190", "193", "390", "184"}


def normalizar(texto: str) -> str:
    """Para comparar nombres de cliente sin acentos, puntos ni mayusculas."""
    if texto is None:
        return ""
    t = unicodedata.normalize("NFD", str(texto).upper())
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return re.sub(r"[^A-Z0-9]", "", t)


@dataclass
class Celda:
    cliente: str
    modelo: str
    estado: str
    marcas: list[str] = field(default_factory=list)
    bruto: str = ""

    @property
    def accionable(self) -> bool:
        return self.estado in ACCIONABLES

    def __str__(self) -> str:
        m = f" [{', '.join(self.marcas)}]" if self.marcas else ""
        return f"{self.estado}{m}"


@dataclass
class Control:
    ruta: Path
    hoja_matriz: str
    modelos: list[str]
    celdas: list[Celda]
    clientes: list[str]
    observaciones: dict[str, str]
    vencimientos: list[dict]
    alquileres: list[dict]

    def filtrar(self, estado=None, modelo=None, cliente=None, con_marca=None) -> list[Celda]:
        salida = self.celdas
        if estado:
            objetivo = [e.strip().lower() for e in estado.split(",")]
            salida = [c for c in salida if c.estado.lower() in objetivo]
        if modelo:
            objetivo = [m.strip().lower() for m in modelo.split(",")]
            salida = [c for c in salida if c.modelo.lower() in objetivo]
        if cliente:
            clave = normalizar(cliente)
            salida = [c for c in salida if clave in normalizar(c.cliente)]
        if con_marca:
            salida = [c for c in salida if con_marca.lower() in [m.lower() for m in c.marcas]]
        return salida


def parsear_celda(valor) -> tuple[str, list[str]]:
    """'Presentado [T, env]' -> ('Presentado', ['T', 'env'])."""
    if valor is None:
        return "", []
    texto = str(valor).strip()
    if not texto:
        return "", []
    m = re.match(r"^([^\[]+?)\s*(?:\[(.*?)\])?\s*$", texto)
    if not m:
        return texto, []
    estado = m.group(1).strip()
    marcas = [x.strip() for x in (m.group(2) or "").split(",") if x.strip()]
    # Normaliza a la nomenclatura oficial del control aunque varie el case.
    for oficial in ESTADOS:
        if normalizar(oficial) == normalizar(estado):
            return oficial, marcas
    return estado, marcas


def localizar_hoja_matriz(wb, preferida: str | None) -> str:
    if preferida and preferida in wb.sheetnames:
        return preferida
    # La matriz es la hoja cuyo nombre encaja con un periodo: 2T-2026, 4T-2025...
    candidatas = [n for n in wb.sheetnames if re.fullmatch(r"[1-4]T-\d{4}", n)]
    if candidatas:
        return sorted(candidatas)[-1]
    raise SystemExit(
        "No encuentro la hoja de la matriz (formato 'NT-AAAA'). "
        f"Hojas disponibles: {', '.join(wb.sheetnames)}. Usa --hoja."
    )


def cargar(ruta: Path, hoja: str | None = None) -> Control:
    import warnings

    warnings.filterwarnings("ignore", category=UserWarning)
    wb = openpyxl.load_workbook(ruta, data_only=True)
    nombre_hoja = localizar_hoja_matriz(wb, hoja)
    ws = wb[nombre_hoja]

    # La fila de cabecera es la que contiene 'NOMBRE' en la primera columna.
    fila_cabecera = None
    for fila in range(1, 12):
        if normalizar(ws.cell(fila, 1).value) == "NOMBRE":
            fila_cabecera = fila
            break
    if fila_cabecera is None:
        raise SystemExit(f"No encuentro la fila de cabecera ('NOMBRE') en la hoja {nombre_hoja}")

    modelos: dict[int, str] = {}
    col_observaciones = None
    for col in range(2, ws.max_column + 1):
        etiqueta = ws.cell(fila_cabecera, col).value
        if etiqueta is None:
            continue
        etiqueta = str(etiqueta).strip()
        if not etiqueta:
            continue
        if re.match(r"^(BANCO|Observaciones|OJO|Alquileres)", etiqueta, re.I):
            if col_observaciones is None:
                col_observaciones = col
            continue
        modelos[col] = etiqueta

    celdas: list[Celda] = []
    clientes: list[str] = []
    observaciones: dict[str, str] = {}

    for fila in range(fila_cabecera + 1, ws.max_row + 1):
        nombre = ws.cell(fila, 1).value
        if not nombre or not str(nombre).strip():
            continue
        nombre = str(nombre).strip()
        # Descarta unicamente las filas de leyenda: su primera celda es el nombre
        # de un estado. Nunca descartes por longitud del nombre: hay clientes
        # cuyo nombre son solo tres letras.
        if normalizar(nombre) in {normalizar(e) for e in ESTADOS}:
            continue
        fila_celdas = []
        for col, modelo in modelos.items():
            estado, marcas = parsear_celda(ws.cell(fila, col).value)
            if estado:
                fila_celdas.append(
                    Celda(nombre, modelo, estado, marcas, str(ws.cell(fila, col).value).strip())
                )
        clientes.append(nombre)
        celdas.extend(fila_celdas)
        if col_observaciones:
            for col in range(col_observaciones, ws.max_column + 1):
                v = ws.cell(fila, col).value
                if v and str(v).strip():
                    observaciones.setdefault(nombre, "")
                    observaciones[nombre] += (" | " if observaciones[nombre] else "") + str(v).strip()

    vencimientos = cargar_vencimientos(wb)
    alquileres = cargar_alquileres(wb)

    return Control(
        ruta=ruta,
        hoja_matriz=nombre_hoja,
        modelos=list(modelos.values()),
        celdas=celdas,
        clientes=clientes,
        observaciones=observaciones,
        vencimientos=vencimientos,
        alquileres=alquileres,
    )


def cargar_vencimientos(wb) -> list[dict]:
    if "Calendario" not in wb.sheetnames:
        return []
    ws = wb["Calendario"]
    # Localiza la cabecera por sus etiquetas en vez de adivinar por tipo de dato:
    # la hoja tiene dos tablas (vencimientos y tareas) con columnas distintas.
    etiquetas = {"CONCEPTO": "concepto", "MODELOS": "modelos", "PERIODO": "periodo",
                 "FECHALIMITE": "fecha_limite", "ESTADO": "estado", "NOTAS": "notas",
                 "TAREA": "concepto", "CLIENTE": "modelos", "RESPONSABLE": "responsable"}
    salida = []
    mapa: dict[int, str] = {}
    for fila in range(1, ws.max_row + 1):
        cabecera = {}
        for col in range(1, ws.max_column + 1):
            clave = etiquetas.get(normalizar(ws.cell(fila, col).value))
            if clave:
                cabecera[col] = clave
        if "concepto" in cabecera.values() and "fecha_limite" in cabecera.values():
            mapa = cabecera
            continue
        if not mapa:
            continue
        registro = {"concepto": "", "modelos": "", "periodo": "", "estado": "", "fecha_limite": None}
        for col, clave in mapa.items():
            valor = ws.cell(fila, col).value
            if valor is None:
                continue
            if clave == "fecha_limite":
                registro["fecha_limite"] = valor.date() if isinstance(valor, datetime) else valor
            else:
                registro[clave] = str(valor).strip()
        if isinstance(registro["fecha_limite"], date) and registro["concepto"]:
            salida.append(registro)
    return salida


def cargar_alquileres(wb) -> list[dict]:
    if "Alquileres" not in wb.sheetnames:
        return []
    ws = wb["Alquileres"]
    meses = ["ENE", "FEB", "MAR", "ABR", "MAY", "JUN", "JUL", "AGO", "SEP", "OCT", "NOV", "DIC"]
    fila_cab = None
    for fila in range(1, 10):
        valores = [normalizar(ws.cell(fila, c).value) for c in range(1, min(20, ws.max_column + 1))]
        if "ENE" in valores and "DIC" in valores:
            fila_cab = fila
            break
    if fila_cab is None:
        return []
    col_mes = {}
    for col in range(1, ws.max_column + 1):
        v = normalizar(ws.cell(fila_cab, col).value)
        if v in meses:
            col_mes[v] = col
    salida = []
    for fila in range(fila_cab + 1, ws.max_row + 1):
        nombre = ws.cell(fila, 1).value
        if not nombre or not str(nombre).strip():
            continue
        estados = {m: (ws.cell(fila, c).value or "") for m, c in col_mes.items()}
        salida.append({"inmueble": str(nombre).strip(), "meses": estados})
    return salida


# --------------------------------------------------------------------------
# Salidas
# --------------------------------------------------------------------------


def pct(parte: int, total: int) -> str:
    return f"{(parte / total * 100):.1f} %" if total else "—"


def cmd_resumen(ctl: Control, args) -> int:
    total = len(ctl.celdas)
    en_flujo = [c for c in ctl.celdas if c.estado in FLUJO]
    presentados = [c for c in ctl.celdas if c.estado in ("Presentado", "Liquidación pendiente")]
    accionables = [c for c in ctl.celdas if c.accionable]

    print(f"CONTROL {ctl.hoja_matriz} — {ctl.ruta.name}")
    print(f"  Clientes:           {len(ctl.clientes)}")
    print(f"  Modelos en matriz:  {len(ctl.modelos)}  ({', '.join(ctl.modelos[:14])}…)")
    print(f"  Celdas con dato:    {total}")
    print()
    print(f"  Presentado:         {len(presentados):4}  ({pct(len(presentados), len(en_flujo))} de lo que esta en flujo)")
    print(f"  Trabajo vivo:       {len(accionables):4}")
    for estado in ACCIONABLES:
        n = sum(1 for c in ctl.celdas if c.estado == estado)
        if n:
            print(f"      {estado:<16}{n:4}")
    for estado in ("Liquidación pendiente", "No aplica", "Baja"):
        n = sum(1 for c in ctl.celdas if c.estado == estado)
        if n:
            print(f"  {estado:<20}{n:4}")

    if accionables:
        print("\n  Trabajo vivo por modelo:")
        por_modelo: dict[str, int] = {}
        for c in accionables:
            por_modelo[c.modelo] = por_modelo.get(c.modelo, 0) + 1
        for modelo, n in sorted(por_modelo.items(), key=lambda x: -x[1])[:12]:
            print(f"      {modelo:<10}{n:4}")

    revisar = [c for c in ctl.celdas if c.estado == "Revisar"]
    if revisar:
        print(f"\n  ATENCION — {len(revisar)} celdas en 'Revisar' (preparadas, esperando revision):")
        for c in revisar[: args.limite]:
            print(f"      {c.cliente[:38]:<38} {c.modelo:<8} {c}")
        if len(revisar) > args.limite:
            print(f"      … y {len(revisar) - args.limite} mas (usa: cola --estado Revisar)")

    hoy = date.today()
    proximos = sorted(
        [v for v in ctl.vencimientos if v["estado"] != "Hecho"], key=lambda v: v["fecha_limite"]
    )
    vencidos = [v for v in proximos if v["fecha_limite"] < hoy]
    futuros = [v for v in proximos if v["fecha_limite"] >= hoy]
    if vencidos:
        print(f"\n  VENCIDOS sin cerrar ({len(vencidos)}):")
        for v in vencidos[:6]:
            print(f"      {v['fecha_limite']}  {v['concepto'][:44]:<44} {v['modelos'][:26]}")
    if futuros:
        print("\n  Proximos vencimientos:")
        for v in futuros[:5]:
            dias = (v["fecha_limite"] - hoy).days
            sem = "●" if dias <= 7 else ("◐" if dias <= 15 else "○")
            print(f"      {sem} {v['fecha_limite']}  ({dias:+4} d)  {v['concepto'][:40]:<40} {v['modelos'][:24]}")
    return 0


def cmd_cola(ctl: Control, args) -> int:
    celdas = ctl.filtrar(args.estado, args.modelo, args.cliente, args.marca)
    if args.solo_accionables:
        celdas = [c for c in celdas if c.accionable]
    orden = {e: i for i, e in enumerate(ACCIONABLES)}
    celdas.sort(key=lambda c: (orden.get(c.estado, 99), c.cliente, c.modelo))

    if args.json:
        print(json.dumps(
            [{"cliente": c.cliente, "modelo": c.modelo, "estado": c.estado, "marcas": c.marcas} for c in celdas],
            ensure_ascii=False, indent=2))
        return 0

    if not celdas:
        print("Sin resultados para ese filtro.")
        return 0

    print(f"{len(celdas)} celdas\n")
    actual = None
    for c in celdas[: args.limite]:
        if c.estado != actual:
            actual = c.estado
            print(f"--- {actual} ---")
        obs = ctl.observaciones.get(c.cliente, "")
        extra = f"   ← {obs[:60]}" if obs and args.observaciones else ""
        print(f"  {c.cliente[:40]:<40} {c.modelo:<9} {str(c):<26}{extra}")
    if len(celdas) > args.limite:
        print(f"\n… y {len(celdas) - args.limite} mas. Usa --limite 0 para verlas todas.")
    return 0


def cmd_cliente(ctl: Control, args) -> int:
    clave = normalizar(args.nombre)
    coincidencias = sorted({c.cliente for c in ctl.celdas if clave in normalizar(c.cliente)})
    if not coincidencias:
        print(f"No hay ningun cliente que contenga '{args.nombre}'.", file=sys.stderr)
        cercanos = [n for n in ctl.clientes if clave[:4] and clave[:4] in normalizar(n)]
        if cercanos:
            print("Quizas: " + ", ".join(cercanos[:8]), file=sys.stderr)
        return 1
    if len(coincidencias) > 1 and not args.primero:
        print(f"{len(coincidencias)} coincidencias:")
        for n in coincidencias:
            print(f"  {n}")
        print("\nPrecisa el nombre o usa --primero.")
        return 1

    nombre = coincidencias[0]
    celdas = [c for c in ctl.celdas if c.cliente == nombre]
    print(f"{nombre}   ({ctl.hoja_matriz})\n")
    vivo = [c for c in celdas if c.accionable]
    hecho = [c for c in celdas if c.estado in ("Presentado", "Liquidación pendiente")]
    fuera = [c for c in celdas if c.estado in FUERA_DE_FLUJO]

    if vivo:
        print("  PENDIENTE DE TRABAJO")
        for c in sorted(vivo, key=lambda c: ACCIONABLES.index(c.estado)):
            print(f"    {c.modelo:<10} {c}")
    if hecho:
        print("\n  CERRADO")
        for c in hecho:
            print(f"    {c.modelo:<10} {c}")
    if fuera:
        print("\n  FUERA DE FLUJO: " + ", ".join(f"{c.modelo} ({c.estado})" for c in fuera))
    obs = ctl.observaciones.get(nombre)
    if obs:
        print(f"\n  OBSERVACIONES\n    {obs}")
    return 0


def cmd_modelo(ctl: Control, args) -> int:
    celdas = [c for c in ctl.celdas if c.modelo.lower() == args.codigo.lower()]
    if not celdas:
        print(f"El modelo '{args.codigo}' no esta en la matriz. Disponibles: {', '.join(ctl.modelos)}")
        return 1
    conteo: dict[str, list[str]] = {}
    for c in celdas:
        conteo.setdefault(c.estado, []).append(c.cliente)
    print(f"Modelo {args.codigo} — {len(celdas)} clientes con dato\n")
    for estado in ESTADOS:
        if estado not in conteo:
            continue
        clientes = sorted(conteo[estado])
        print(f"  {estado:<22}{len(clientes):4}")
        if estado in ACCIONABLES:
            for n in clientes[: args.limite]:
                print(f"        {n}")
            if len(clientes) > args.limite:
                print(f"        … y {len(clientes) - args.limite} mas")
    return 0


def cmd_vencimientos(ctl: Control, args) -> int:
    hoy = date.today()
    filas = sorted(ctl.vencimientos, key=lambda v: v["fecha_limite"])
    if not filas:
        print("El control no tiene hoja 'Calendario' legible.")
        return 1
    print(f"Hoy: {hoy.strftime('%d/%m/%Y')}\n")
    for v in filas:
        dias = (v["fecha_limite"] - hoy).days
        if args.dias is not None and dias > args.dias:
            continue
        if v["estado"] == "Hecho" and not args.todos:
            continue
        if dias < 0:
            sem = "● VENCIDO"
        elif dias <= 7:
            sem = "● ≤7 d   "
        elif dias <= 15:
            sem = "◐ ≤15 d  "
        else:
            sem = "○ en plazo"
        print(f"  {sem}  {v['fecha_limite']}  ({dias:+5} d)  {v['concepto'][:38]:<38} {v['modelos'][:30]}")
    return 0


def cmd_alquileres(ctl: Control, args) -> int:
    if not ctl.alquileres:
        print("El control no tiene hoja 'Alquileres' legible.")
        return 1
    meses = ["ENE", "FEB", "MAR", "ABR", "MAY", "JUN", "JUL", "AGO", "SEP", "OCT", "NOV", "DIC"]
    limite = args.hasta_mes or date.today().month
    print(f"Alquileres — revisando hasta {meses[limite - 1]}\n")
    incidencias = 0
    for fila in ctl.alquileres:
        pendientes = [
            m for i, m in enumerate(meses[:limite])
            if str(fila["meses"].get(m, "")).strip() in ("Pendiente", "Parcial", "Revisar", "Emitido")
        ]
        if pendientes:
            incidencias += 1
            estados = {m: str(fila["meses"].get(m, "")).strip() for m in pendientes}
            print(f"  {fila['inmueble'][:44]:<44} {', '.join(f'{m}:{e}' for m, e in estados.items())}")
    if not incidencias:
        print("  Todo facturado hasta la fecha.")
    else:
        print(f"\n  {incidencias} inmuebles con meses sin cerrar.")
    return 0


def cmd_huecos(ctl: Control, args) -> int:
    """Audita el propio control: lo que falta o es incoherente."""
    problemas: list[tuple[str, str]] = []

    for cliente in ctl.clientes:
        celdas = [c for c in ctl.celdas if c.cliente == cliente]
        if not celdas:
            problemas.append(("SIN DATOS", f"{cliente}: ninguna celda cumplimentada"))
            continue
        sin_dato = [c.modelo for c in celdas if c.estado == "Sin dato"]
        if sin_dato:
            problemas.append(("SIN DATO", f"{cliente}: {', '.join(sin_dato)}"))
        # Un cliente con 303 en flujo deberia tener 390 previsto, y viceversa.
        mapa = {c.modelo: c.estado for c in celdas}
        for periodico, anual in (("303", "390"), ("111", "190"), ("115", "180"), ("123", "193")):
            if mapa.get(periodico) in FLUJO and mapa.get(anual) in FUERA_DE_FLUJO:
                problemas.append((
                    "COHERENCIA",
                    f"{cliente}: {periodico} en flujo pero {anual} marcado '{mapa[anual]}'",
                ))
        # Sociedad con 200 pero sin 202, o al reves.
        if mapa.get("200/24") in FLUJO and mapa.get("202") == "No aplica":
            problemas.append(("COHERENCIA", f"{cliente}: presenta 200 pero el 202 esta 'No aplica' (revisar art. 40.2 LIS)"))

    faltan_modelos = [m for m in ("111", "303", "347", "390") if m not in ctl.modelos]
    if faltan_modelos:
        problemas.append(("ESTRUCTURA", f"La matriz no tiene columna para: {', '.join(faltan_modelos)}"))

    if not problemas:
        print("Sin incidencias detectadas en el control.")
        return 0

    print(f"{len(problemas)} incidencias\n")
    actual = None
    for tipo, texto in sorted(problemas):
        if tipo != actual:
            actual = tipo
            print(f"--- {tipo} ---")
        print(f"  {texto}")
    print("\nEsto audita la COHERENCIA del control, no la correccion fiscal.")
    return 1


def cmd_marcar(ctl: Control, args) -> int:
    import warnings

    warnings.filterwarnings("ignore", category=UserWarning)
    if args.estado not in ESTADOS:
        print(f"Estado no valido. Admitidos: {', '.join(ESTADOS)}", file=sys.stderr)
        return 2

    wb = openpyxl.load_workbook(ctl.ruta)
    ws = wb[ctl.hoja_matriz]
    fila_cab = next(
        f for f in range(1, 12) if normalizar(ws.cell(f, 1).value) == "NOMBRE"
    )
    col = next(
        (c for c in range(2, ws.max_column + 1)
         if str(ws.cell(fila_cab, c).value or "").strip().lower() == args.modelo.lower()),
        None,
    )
    if col is None:
        print(f"No encuentro la columna del modelo '{args.modelo}'.", file=sys.stderr)
        return 1

    clave = normalizar(args.cliente)
    filas = [
        f for f in range(fila_cab + 1, ws.max_row + 1)
        if ws.cell(f, 1).value and clave in normalizar(ws.cell(f, 1).value)
    ]
    if len(filas) != 1:
        print(
            f"{'Ningun' if not filas else len(filas)} cliente coincide con '{args.cliente}'."
            + (" Precisa el nombre." if filas else ""),
            file=sys.stderr,
        )
        for f in filas[:8]:
            print(f"  {ws.cell(f, 1).value}", file=sys.stderr)
        return 1

    fila = filas[0]
    anterior = ws.cell(fila, col).value
    nuevo = args.estado + (f" [{', '.join(args.marca)}]" if args.marca else "")
    if args.simular:
        print(f"SIMULACION — {ws.cell(fila, 1).value} / {args.modelo}: '{anterior}' -> '{nuevo}'")
        return 0
    ws.cell(fila, col).value = nuevo
    destino = Path(args.salida) if args.salida else ctl.ruta
    wb.save(destino)
    print(f"{ws.cell(fila, 1).value} / {args.modelo}: '{anterior}' -> '{nuevo}'")
    print(f"Guardado en {destino}")
    if destino == ctl.ruta:
        print("Aviso: se ha sobrescrito el control. Los formatos condicionales se conservan,")
        print("pero revisa el fichero antes de compartirlo.")
    return 0


def cmd_exportar(ctl: Control, args) -> int:
    filas = []
    for cliente in ctl.clientes:
        fila = {"cliente": cliente, "observaciones": ctl.observaciones.get(cliente, "")}
        for c in ctl.celdas:
            if c.cliente == cliente:
                fila[c.modelo] = c.bruto
        filas.append(fila)
    if args.formato == "json":
        print(json.dumps(filas, ensure_ascii=False, indent=2))
    else:
        import csv

        columnas = ["cliente"] + ctl.modelos + ["observaciones"]
        escritor = csv.DictWriter(sys.stdout, fieldnames=columnas, delimiter=";", extrasaction="ignore")
        escritor.writeheader()
        escritor.writerows(filas)
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--fichero", type=Path, default=Path("Control.xlsx"))
    ap.add_argument("--hoja", help="Hoja de la matriz (por defecto, la del periodo mas reciente)")
    sub = ap.add_subparsers(dest="comando", required=True)

    p = sub.add_parser("resumen"); p.add_argument("--limite", type=int, default=12)
    p = sub.add_parser("cola")
    p.add_argument("--estado", help="Uno o varios separados por coma")
    p.add_argument("--modelo"); p.add_argument("--cliente"); p.add_argument("--marca")
    p.add_argument("--solo-accionables", action="store_true")
    p.add_argument("--observaciones", action="store_true")
    p.add_argument("--limite", type=int, default=60); p.add_argument("--json", action="store_true")
    p = sub.add_parser("cliente"); p.add_argument("nombre"); p.add_argument("--primero", action="store_true")
    p = sub.add_parser("modelo"); p.add_argument("codigo"); p.add_argument("--limite", type=int, default=40)
    p = sub.add_parser("vencimientos"); p.add_argument("--dias", type=int); p.add_argument("--todos", action="store_true")
    p = sub.add_parser("alquileres"); p.add_argument("--hasta-mes", type=int)
    sub.add_parser("huecos")
    p = sub.add_parser("marcar")
    p.add_argument("--cliente", required=True); p.add_argument("--modelo", required=True)
    p.add_argument("--estado", required=True); p.add_argument("--marca", action="append", default=[])
    p.add_argument("--salida"); p.add_argument("--simular", action="store_true")
    p = sub.add_parser("exportar"); p.add_argument("--formato", choices=("csv", "json"), default="csv")

    args = ap.parse_args()
    if not args.fichero.exists():
        print(f"No existe {args.fichero}. Indica la ruta con --fichero.", file=sys.stderr)
        return 2

    ctl = cargar(args.fichero, args.hoja)
    return {
        "resumen": cmd_resumen, "cola": cmd_cola, "cliente": cmd_cliente,
        "modelo": cmd_modelo, "vencimientos": cmd_vencimientos,
        "alquileres": cmd_alquileres, "huecos": cmd_huecos,
        "marcar": cmd_marcar, "exportar": cmd_exportar,
    }[args.comando](ctl, args)


if __name__ == "__main__":
    raise SystemExit(main())
